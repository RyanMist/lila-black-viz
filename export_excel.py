# -*- coding: utf-8 -*-
"""
Generates player_data_verification.xlsx with 6 sheets for data verification.
Run from inside the player_data folder:
    python export_excel.py
"""

import os
import pyarrow.parquet as pq
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

# ── colours ────────────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", fgColor="1A1A2E")
HDR_FONT   = Font(color="FFFFFF", bold=True, size=11)
ALT_FILL   = PatternFill("solid", fgColor="F4F5F7")
TITLE_FONT = Font(bold=True, size=13)
THIN       = Side(style="thin", color="CCCCCC")
BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DATE_FOLDERS = {
    "Feb 10": "February_10",
    "Feb 11": "February_11",
    "Feb 12": "February_12",
    "Feb 13": "February_13",
    "Feb 14": "February_14",
}

MAP_CONFIG = {
    "AmbroseValley": {"scale": 900,  "origin_x": -370, "origin_z": -473},
    "GrandRift":     {"scale": 581,  "origin_x": -290, "origin_z": -290},
    "Lockdown":      {"scale": 1000, "origin_x": -500, "origin_z": -500},
}

# ─────────────────────────────────────────────────────────────
# 1. LOAD ALL DATA
# ─────────────────────────────────────────────────────────────
print("Loading parquet files...")
frames = []
file_count = 0
for date_label, folder in DATE_FOLDERS.items():
    folder_path = os.path.join(os.path.dirname(__file__), folder)
    if not os.path.isdir(folder_path):
        continue
    for fname in os.listdir(folder_path):
        fpath = os.path.join(folder_path, fname)
        try:
            df = pq.read_table(fpath).to_pandas()
            df["event"] = df["event"].apply(
                lambda e: e.decode("utf-8") if isinstance(e, bytes) else e
            )
            df["date"] = date_label
            df["filename"] = fname
            frames.append(df)
            file_count += 1
        except Exception:
            continue

df_all = pd.concat(frames, ignore_index=True)
df_all["is_human"] = ~df_all["user_id"].str.match(r"^\d+$", na=False)
df_all["player_type"] = df_all["is_human"].map({True: "Human", False: "Bot"})
df_all["match_id_clean"] = df_all["match_id"].str.replace(".nakama-0", "", regex=False)
print(f"  Loaded {file_count:,} files, {len(df_all):,} rows.")


# ─────────────────────────────────────────────────────────────
# 2. BUILD SHEETS
# ─────────────────────────────────────────────────────────────

# ── Sheet 1: Overview ────────────────────────────────────────
overview_rows = [
    ["Metric", "Value"],
    ["Date range", "Feb 10 – Feb 14, 2026"],
    ["Total files loaded", file_count],
    ["Total event rows", len(df_all)],
    ["Unique human players", df_all[df_all["is_human"]]["user_id"].nunique()],
    ["Unique bots", df_all[~df_all["is_human"]]["user_id"].nunique()],
    ["Unique matches", df_all["match_id_clean"].nunique()],
    ["Maps", ", ".join(sorted(df_all["map_id"].unique()))],
    [],
    ["Events by type", "Count", "% of total"],
]
for evt, cnt in df_all["event"].value_counts().items():
    overview_rows.append([evt, cnt, round(cnt / len(df_all) * 100, 2)])

overview_rows += [
    [],
    ["Files per day", "Files", "Rows"],
]
for date_label in DATE_FOLDERS:
    sub = df_all[df_all["date"] == date_label]
    overview_rows.append([date_label, sub["filename"].nunique(), len(sub)])

df_overview = pd.DataFrame(overview_rows)

# ── Sheet 2: Match Summary ───────────────────────────────────
match_rows = []
for match_id, g in df_all.groupby("match_id_clean"):
    humans  = g[g["is_human"]]
    bots    = g[~g["is_human"]]
    dur_s   = (g["ts"].max() - g["ts"].min()).total_seconds()
    mins, secs = divmod(int(dur_s), 60)
    match_rows.append({
        "Match ID":        match_id,
        "Map":             g["map_id"].iloc[0],
        "Date":            g["date"].iloc[0],
        "Human Players":   humans["user_id"].nunique(),
        "Bots":            bots["user_id"].nunique(),
        "Total Events":    len(g),
        "Kills (H>H)":     (g["event"] == "Kill").sum(),
        "Deaths (H)":      (g["event"] == "Killed").sum(),
        "Bot Kills":       (g["event"] == "BotKill").sum(),
        "Bot Deaths":      (g["event"] == "BotKilled").sum(),
        "Storm Deaths":    (g["event"] == "KilledByStorm").sum(),
        "Loot Events":     (g["event"] == "Loot").sum(),
        "Duration":        f"{mins}m {secs:02d}s",
        "Duration (s)":    round(dur_s),
    })
df_matches = pd.DataFrame(match_rows).sort_values(["Date", "Match ID"])

# ── Sheet 3: Player Summary ──────────────────────────────────
player_rows = []
for uid, g in df_all[df_all["is_human"]].groupby("user_id"):
    matches_played = g["match_id_clean"].nunique()
    player_rows.append({
        "Player ID":       uid,
        "Matches Played":  matches_played,
        "Maps Played":     ", ".join(sorted(g["map_id"].unique())),
        "Kills":           (g["event"] == "Kill").sum(),
        "Deaths":          (g["event"].isin(["Killed", "KilledByStorm"])).sum(),
        "Bot Kills":       (g["event"] == "BotKill").sum(),
        "Loot Events":     (g["event"] == "Loot").sum(),
        "Storm Deaths":    (g["event"] == "KilledByStorm").sum(),
        "K/D Ratio":       round(
            (g["event"] == "Kill").sum() /
            max((g["event"].isin(["Killed", "KilledByStorm"])).sum(), 1), 2
        ),
        "Dates Active":    ", ".join(sorted(g["date"].unique(), key=lambda d: list(DATE_FOLDERS).index(d))),
    })
df_players = pd.DataFrame(player_rows).sort_values("Kills", ascending=False)

# ── Sheet 4: Events by Map & Date ────────────────────────────
pivot = (
    df_all.groupby(["map_id", "date", "event"])
    .size()
    .reset_index(name="count")
    .pivot_table(index=["map_id", "date"], columns="event", values="count", fill_value=0)
    .reset_index()
)
pivot.columns.name = None
df_map_date = pivot.rename(columns={"map_id": "Map", "date": "Date"})

# ── Sheet 5: Raw Sample (500 rows per map) ───────────────────
samples = []
for map_id in sorted(df_all["map_id"].unique()):
    sub = df_all[df_all["map_id"] == map_id].sample(
        n=min(500, len(df_all[df_all["map_id"] == map_id])),
        random_state=42
    )
    samples.append(sub)
df_sample = pd.concat(samples, ignore_index=True)[[
    "date", "map_id", "player_type", "user_id",
    "match_id_clean", "event", "x", "y", "z", "ts"
]].rename(columns={
    "map_id": "map", "match_id_clean": "match_id",
    "player_type": "type"
}).sort_values(["date", "map", "match_id", "ts"])

# ── Sheet 6: Coordinate Spot-check ──────────────────────────
coord_rows = []
for map_id, cfg in MAP_CONFIG.items():
    sub = df_all[
        (df_all["map_id"] == map_id) &
        (~df_all["pixel_x"].isna() if "pixel_x" in df_all.columns else True)
    ].head(100) if "pixel_x" in df_all.columns else df_all[df_all["map_id"] == map_id].head(100)
    for _, row in sub.iterrows():
        x, z = float(row["x"]), float(row["z"])
        u = (x - cfg["origin_x"]) / cfg["scale"]
        v = (z - cfg["origin_z"]) / cfg["scale"]
        px = round(u * 1024, 1)
        py = round((1 - v) * 1024, 1)
        in_bounds = 0 <= px <= 1024 and 0 <= py <= 1024
        coord_rows.append({
            "Map":          map_id,
            "Event":        row["event"],
            "World X":      round(x, 2),
            "World Z":      round(z, 2),
            "U (0-1)":      round(u, 4),
            "V (0-1)":      round(v, 4),
            "Pixel X":      px,
            "Pixel Y":      py,
            "In Bounds?":   "YES" if in_bounds else "NO",
        })
df_coords = pd.DataFrame(coord_rows)


# ─────────────────────────────────────────────────────────────
# 3. WRITE TO EXCEL
# ─────────────────────────────────────────────────────────────
out_path = os.path.join(os.path.dirname(__file__), "player_data_verification.xlsx")
print(f"Writing {out_path} ...")

with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
    # Write raw data first (no header for overview since it's manual)
    pd.DataFrame(df_overview).to_excel(writer, sheet_name="Overview",      index=False, header=False)
    df_matches.to_excel(              writer, sheet_name="Match Summary",   index=False)
    df_players.to_excel(              writer, sheet_name="Player Summary",  index=False)
    df_map_date.to_excel(             writer, sheet_name="Events by Map+Date", index=False)
    df_sample.to_excel(               writer, sheet_name="Raw Sample",      index=False)
    df_coords.to_excel(               writer, sheet_name="Coord Check",     index=False)

# ─────────────────────────────────────────────────────────────
# 4. STYLE THE WORKBOOK
# ─────────────────────────────────────────────────────────────
wb = load_workbook(out_path)

def style_sheet(ws, header_row=1):
    """Apply header style, alternating rows, auto column widths, freeze pane."""
    # Header row
    for cell in ws[header_row]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    # Data rows – alternating fill + border
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1), start=1):
        fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
        for cell in row:
            cell.fill = fill
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center")

    # Auto column widths (capped at 50)
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=10
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 50)

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.row_dimensions[header_row].height = 28


# ── Overview sheet (manual header rows, not a standard table) ─
ws_ov = wb["Overview"]
# Style the two section headers manually
section_rows = [1, 10, 19]
for r in section_rows:
    cell = ws_ov.cell(row=r, column=1)
    cell.font = Font(bold=True, size=12, color="1A1A2E")
# Style first row as title
title_cell = ws_ov.cell(row=1, column=1)
title_cell.value = "LILA BLACK  —  Data Verification Summary"
title_cell.font = Font(bold=True, size=14, color="1A1A2E")
ws_ov.merge_cells("A1:C1")
# Auto-width overview cols
for col in ws_ov.columns:
    max_len = max(
        (len(str(c.value)) for c in col if c.value is not None), default=10
    )
    ws_ov.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 55)

# ── Style remaining sheets ─────────────────────────────────
for sheet_name in ["Match Summary", "Player Summary", "Events by Map+Date", "Raw Sample", "Coord Check"]:
    style_sheet(wb[sheet_name])

# ── Colour scale on Match Summary "Duration (s)" column ──────
ws_ms = wb["Match Summary"]
dur_col = None
for cell in ws_ms[1]:
    if cell.value == "Duration (s)":
        dur_col = get_column_letter(cell.column)
        break
if dur_col:
    last_row = ws_ms.max_row
    ws_ms.conditional_formatting.add(
        f"{dur_col}2:{dur_col}{last_row}",
        ColorScaleRule(
            start_type="min", start_color="FFF5F5",
            mid_type="percentile", mid_value=50, mid_color="FFD700",
            end_type="max", end_color="E03B3B",
        ),
    )

# ── YES/NO colour in Coord Check sheet ────────────────────────
ws_cc = wb["Coord Check"]
for row in ws_cc.iter_rows(min_row=2):
    for cell in row:
        if cell.value == "NO":
            cell.fill = PatternFill("solid", fgColor="FFCCCC")
            cell.font = Font(bold=True, color="CC0000")
        elif cell.value == "YES":
            cell.fill = PatternFill("solid", fgColor="CCFFCC")
            cell.font = Font(color="006600")

# ── Tab colours ───────────────────────────────────────────────
tab_colours = {
    "Overview":          "1A1A2E",
    "Match Summary":     "E03B3B",
    "Player Summary":    "4A90D9",
    "Events by Map+Date":"FF8C00",
    "Raw Sample":        "6C757D",
    "Coord Check":       "9B59B6",
}
for sheet_name, colour in tab_colours.items():
    if sheet_name in wb.sheetnames:
        wb[sheet_name].sheet_properties.tabColor = colour

wb.save(out_path)
print(f"Done. Saved to: {out_path}")
print(f"\nSheets:")
for s in wb.sheetnames:
    ws = wb[s]
    print(f"  {s:25s}  {ws.max_row-1:>6,} data rows")
